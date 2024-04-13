import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
import win32com.client
import pypdf
import os
import glob
import datetime
import shutil

def get_file_information(fpath2):
    """月例点検結果データ、点検基準表のファイル名をファイル名情報.xlsxから読込み
    Returns:
    -------
    inspection_data: str
        月例点検結果データ
  　base_inspection: str
       点検基準表データ
    """
    path = f"{f_path2}/ファイル名情報.xlsx"
    wb = load_workbook(path)
    ws = wb['Sheet1'] 
    
    inspection_data = ws["C2"].value
    base_inspection = ws["C3"].value
    
    wb.close()

    return inspection_data, base_inspection
def input_excel(new_form,machine_name,kiki_n,room_name,bikou_data,f_path,template_path):
    """データフレームをエクセルファイル（月例点検結果_雛型）に入力
       生成したexcelは「tmp_data_file」内に保存
    Parameter:
    ---------
    new_form: dataframe
    machine_name: str
    kiki_n: str
    room_name: str
        機器設置部屋
    bikou: str
        点検結果の備考
    f_path: str
        現在位置のパス
    template_path: str
        雛型エクセルデータのパス
    Return:
    ------
    folder_path: str
        生成したExcelデータを保存したフォルダ(tmp_data_file)の絶対パス
    """

    list_d = [list(new_form.loc[i,:]) for i in range(len(new_form))]

    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    
    wb = load_workbook(template_path)
    ws = wb['Sheet1'] 

    ws.cell(row=4,column=3,value=machine_name)
    ws.cell(row=5,column=3,value=kiki_n)
    ws.cell(row=6,column=3,value=room_name)
    ws.cell(row=7,column=7,value=bikou_data)

    for y, row in enumerate(list_d):
        for x, cells in enumerate(row):
            ws.cell(row=12+y,column=2+x,value=list_d[y][x])
            if ws.cell(row=12+y,column=2+x).value:
                ws.cell(row=12+y,column=2+x).border= border           

    folder_path = f"{f_path}/tmp_data_file"
    
    os.makedirs(folder_path,exist_ok="true")

    wb.save(f'{f_path}/tmp_data_file/{kiki_n}.xlsx')

    return folder_path

def make_df(kijyun,kiki_n):
    """点検結果統合表から指定された機器番号に該当する点検データを抽出し、空データフレームを作成
    Parameter
    ---------
    kijyun: dataframe
            点検基準表
    kiki_n：str
            base_dataに記載の機器番号
    Returns
    -------
    new_form_: dataframe
    new_form: dataframe
    """
    ex_room = kijyun[kijyun["該当機器"]==kiki_n].reset_index()
    new_form_ = ex_room.loc[:,["点検番号","点検部位","点検内容","点検方法","判定基準"]]
    new_form_[["4月","5月","6月","7月","8月","9月","10月","11月","12月","1月","2月","3月"]]=""
    add_data = pd.DataFrame({"点検番号":"","点検部位":"","点検内容":"","点検方法":"","判定基準":"点検者"},index=[len(new_form_)])
    add_data2 = pd.DataFrame({"点検番号":"","点検部位":"","点検内容":"","点検方法":"","判定基準":"安全衛生委員"},index=[len(new_form_)+1])
    add_data3 = pd.DataFrame({"点検番号":"","点検部位":"","点検内容":"","点検方法":"","判定基準":"室長"},index=[len(new_form_)+2])
    new_form = pd.concat([new_form_,add_data,add_data2,add_data3])

    return new_form_,new_form

def extract_machine_room(kiki_n,result):
    """月例点検結果データから機器番号に該当する機器名、設置部屋情報を取得
    Parmeter:
    --------
    kiki_n: str
        機器番号
    result: dataframe
        月例点検結果のデータフレーム
    Returns:
    -------
    machine_name: str
        機器名
    room_name: str
        設置部屋名
    """
    machine_name = list(result[result['機器番号']==kiki_n]["装置名"])[0]
    room_name = list(result[result['機器番号']==kiki_n]["設置場所"])[0]

    return machine_name,room_name

def input_result(result,kiki_n,new_form_,new_form):
    """機器毎に生成した空データフレームに点検結果を入力
    Prameter
    --------
    result: dataframe
            点検結果データ（全ての機器、月の点検データを含む）
    kiki_n: str
            書き込みを行う機器番号（点検結果データから抽出）
    new_form_: dataframe
            機器毎にmake_dfで生成した空データフレーム（点検者、安全衛生、室長カラムを含まない）
    new_form: dataframe
            機器毎にmake_dfで生成した空データフレーム（点検者、安全衛生、室長カラムを含む）            
    Returns
    -------
    new_form: dataframe
    bikou_data: str
    """
    
    bikou_list = []

    for i in range(len(result)):
        if result.loc[i,"機器番号"] == kiki_n:
            tenken_tuki = result.loc[i,"点検月"]
            tenken_man_ = result.loc[i,"点検者"]
            tenken_man = tenken_man_.split(" ")[0]
            anzen = result.loc[i,"安全衛生委員"]
            manager = result.loc[i,"室長"]
            bikou = result.loc[i,"点検結果備考"]

            if bikou:
                bikou2 = f"({tenken_tuki}):{bikou}"
                bikou_list.append(bikou2)
    
            tenken_length = len(new_form_)
            tenken_index = len(new_form_)
            confirm_index = len(new_form_)+1
            approve_index = len(new_form_)+2 
    
            for j in range(tenken_length):
                new_form.loc[j,tenken_tuki]="✓"
    
            new_form.loc[tenken_index,tenken_tuki] = tenken_man
            new_form.loc[confirm_index,tenken_tuki] = anzen
            new_form.loc[approve_index,tenken_tuki] = manager

    bikou_data = "->".join(bikou_list)

    return new_form, bikou_data

def convert_to_pdf(folder_path):
    """フォルダ内のエクセルデータをpdfに変換
    Parameter:
    ---------
    folder_path: str
        エクセルデータを含むフォルダの絶対パス
    """
    excel = win32com.client.Dispatch("Excel.Application")

    for path in glob.glob(f"{folder_path}/*.xlsx"):
        file_name = os.path.basename(path)
        pdf_name = os.path.splitext(os.path.basename(path))[0]

        file_path = f"{folder_path}/{file_name}"
        pdf_path = f"{folder_path}/{pdf_name}"

        file = excel.Workbooks.Open(file_path)
        file.WorkSheets(1).Activate()
        file.ActiveSheet.ExportAsFixedFormat(0,pdf_path)

        file.Close()
        excel.Quit()

def merge_pdf(folder_path,save_path,name):
    """フォルダー内のpdfデータを統合
    Parameter:
    ---------
    folder_path: str
        pdfデータが保管されたフォルダーの絶対パス
    save_path: str
        PDF統合データを保存するフォルダーの絶対パス
    name: str
        PDF統合データの保存ファイル名
    """
    merger = pypdf.PdfWriter()
    
    pdffiles = glob.glob(f"{folder_path}/*.pdf")
    pdffiles.sort()
    
    for p_name in pdffiles:
        merger.append(p_name)
    
    merger.write(f"{save_path}/{name}.pdf")
    merger.close()
     
# 現在値パスの取得
f_path = os.getcwd()
f_path2 = f"{f_path}/ファイル名の登録"
template_path = f"{f_path}/雛型データ/月例点検結果_雛型.xlsx"

# ファイル名の取得
inspection_data, base_inspection = get_file_information(f_path2)
tenken_path = f"{f_path}/{inspection_data}"
kijyun_path = f"{f_path}/{base_inspection}"

# 点検結果の読込み
df = pd.read_csv(tenken_path)
result_ = df.loc[:,["機器番号","装置名","設置場所","点検月","点検者","安全衛生委員","室長","点検結果備考","点検結果"]]
result = result_.fillna("")

# 点検基準表の読込み
kijyun_ = pd.read_csv(kijyun_path)
kijyun = kijyun_.loc[:,["該当機器","点検番号","点検部位","点検内容","点検方法","判定基準"]]

# 点検結果に記載された機器番号リストの作成
kiki_list_ = [i for i in result['機器番号']]
kiki_list = list(set(kiki_list_))

# 機器番号毎に空データフレームを作成、月毎に点検結果を追記、点検者名/安全衛生委員/室長名を追記
for kiki_n in kiki_list:
    new_form_,new_form = make_df(kijyun,kiki_n)
    machine_name,room_name = extract_machine_room(kiki_n,result)
    new_form, bikou_data = input_result(result,kiki_n,new_form_,new_form)
    folder_path = input_excel(new_form,machine_name,kiki_n,room_name,bikou_data,f_path,template_path)

# 生成した月例点検結果(.xlsx)をpdfに変換
convert_to_pdf(folder_path)

# 生成した機器毎のpdfを統合
save_path = os.getcwd()
year = datetime.date.today().year
name = f"{year}年度_月例点検データ"
merge_pdf(folder_path,save_path,name)

# 不要データの削除
shutil.rmtree(folder_path)